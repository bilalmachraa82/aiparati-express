#!/usr/bin/env python3
"""
Teste Real do AutoFund AI com IES PDF
Processa o ficheiro IES - 2023.pdf com Claude API
"""

import os
import sys
import json
from datetime import datetime
from pathlib import Path
import anthropic
from dotenv import load_dotenv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# Carregar ambiente
load_dotenv()

# Configurações
ANTHROPIC_API_KEY = os.getenv('ANTHROPIC_API_KEY')
IES_PATH = "IES - 2023.pdf"
TEMPLATE_PATH = "template_iapmei.xlsx"
OUTPUT_DIR = Path("outputs")
OUTPUT_DIR.mkdir(exist_ok=True)

def test_pdf_extraction():
    """Testa extração do IES PDF com Claude API"""
    print("🚀 Iniciando teste real do AutoFund AI")
    print(f"📄 Ficheiro IES: {IES_PATH}")
    print(f"🔑 API Key: {'✅ Configurada' if ANTHROPIC_API_KEY else '❌ Não configurada'}")

    if not ANTHROPIC_API_KEY:
        print("❌ Configure ANTHROPIC_API_KEY no .env")
        return None

    if not os.path.exists(IES_PATH):
        print(f"❌ Ficheiro IES não encontrado: {IES_PATH}")
        return None

    try:
        # Inicializar cliente Claude
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        print("✅ Cliente Claude inicializado")

        # Ler PDF
        with open(IES_PATH, "rb") as f:
            pdf_data = f.read()

        print(f"📊 PDF lido: {len(pdf_data):,} bytes")

        # Extrair dados com Claude 3.5 Sonnet
        print("\n🧠 Extraindo dados com Claude 3.5 Sonnet...")

        message = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=4000,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": pdf_data.hex()
                        }
                    },
                    {
                        "type": "text",
                        "text": """Extraia os dados financeiros deste IES (Informação Empresarial Simplificada) português.

Retorne APENAS um JSON com esta estrutura:
{
  "nif": "NIF da empresa",
  "designacao_social": "Nome da empresa",
  "ano_exercicio": "Ano do exercício",
  "ativo_total": 123456.78,
  "passivo_total": 123456.78,
  "capital_proprio": 123456.78,
  "volume_negocios": 123456.78,
  "resultados_antes_impostos": 123456.78,
  "depreciacoes_amortizacoes": 123456.78,
  "ebitda": 123456.78
}

Use valores numéricos com 2 casas decimais. Se não encontrar algum valor, use 0.0."""
                    }
                ]
            }]
        )

        # Processar resposta
        content = message.content[0].text if message.content else ""
        print(f"📝 Resposta Claude: {len(content)} caracteres")

        # Extrair JSON
        try:
            # Procurar JSON no conteúdo
            start = content.find("{")
            end = content.rfind("}") + 1
            json_str = content[start:end]
            dados = json.loads(json_str)

            print("\n✅ Dados extraídos com sucesso:")
            print(f"   Empresa: {dados.get('designacao_social', 'N/A')}")
            print(f"   NIF: {dados.get('nif', 'N/A')}")
            print(f"   Ano: {dados.get('ano_exercicio', 'N/A')}")
            print(f"   Ativo Total: €{dados.get('ativo_total', 0):,.2f}")
            print(f"   Passivo Total: €{dados.get('passivo_total', 0):,.2f}")
            print(f"   Capital Próprio: €{dados.get('capital_proprio', 0):,.2f}")
            print(f"   Volume Negócios: €{dados.get('volume_negocios', 0):,.2f}")
            print(f"   EBITDA: €{dados.get('ebitda', 0):,.2f}")

            # Validar equação contabilística
            ativo = dados.get('ativo_total', 0)
            passivo = dados.get('passivo_total', 0)
            capital = dados.get('capital_proprio', 0)
            soma = passivo + capital

            print(f"\n🧮 Validação Contabilística:")
            print(f"   Ativo: €{ativo:,.2f}")
            print(f"   Passivo + CP: €{soma:,.2f}")
            print(f"   Diferença: €{abs(ativo - soma):,.2f}")

            if abs(ativo - soma) < 0.01:
                print("   ✅ Equação contabilística bate!")
            else:
                print("   ⚠️ Equação não bate (diferença > 0.01€)")

            return dados

        except json.JSONDecodeError as e:
            print(f"❌ Erro ao parsear JSON: {e}")
            print(f"   Conteúdo: {content[:500]}...")
            return None

    except Exception as e:
        print(f"❌ Erro na extração: {e}")
        import traceback
        traceback.print_exc()
        return None

def test_excel_generation(dados):
    """Testa geração do Excel IAPMEI"""
    if not dados or not os.path.exists(TEMPLATE_PATH):
        print("❌ Dados ou template não disponíveis")
        return

    print("\n📊 Gerando Excel IAPMEI...")

    try:
        # Carregar template
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        print(f"✅ Template carregado: {TEMPLATE_PATH}")

        # Mapear dados para células
        mapeamento = {
            "A5": dados.get('designacao_social', ''),
            "B10": dados.get('nif', ''),
            "F8": dados.get('ano_exercicio', ''),
            "D25": dados.get('ativo_total', 0),
            "D30": dados.get('passivo_total', 0),
            "D35": dados.get('capital_proprio', 0),
            "D40": dados.get('volume_negocios', 0),
            "D45": dados.get('ebitda', 0)
        }

        # Preencher células
        preenchidas = 0
        for celula, valor in mapeamento.items():
            if valor:
                coord = coordinate_from_string(celula)
                ws.cell(row=coord[1], column=column_index_from_string(coord[0]), value=valor)
                preenchidas += 1

        print(f"✅ {preenchidas} células preenchidas")

        # Salvar arquivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = OUTPUT_DIR / f"autofund_test_{timestamp}.xlsx"
        wb.save(output_file)

        print(f"✅ Excel salvo: {output_file}")
        print(f"   Tamanho: {output_file.stat().st_size:,} bytes")

        # Verificar conteúdo
        wb_check = load_workbook(output_file)
        ws_check = wb_check.active

        print("\n📋 Valores escritos:")
        for celula in ["A5", "B10", "F8", "D25", "D30", "D35", "D40", "D45"]:
            coord = coordinate_from_string(celula)
            valor = ws_check.cell(row=coord[1], column=column_index_from_string(coord[0])).value
            if valor:
                print(f"   {celula}: {valor}")

        return str(output_file)

    except Exception as e:
        print(f"❌ Erro ao gerar Excel: {e}")
        import traceback
        traceback.print_exc()
        return None

def test_financial_analysis(dados):
    """Testa análise financeira com Claude Opus"""
    if not dados:
        return

    print("\n🔬 Gerando análise financeira com Claude Opus...")

    try:
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

        # Calcular métricas
        ativo = dados.get('ativo_total', 0)
        capital = dados.get('capital_proprio', 0)
        volume = dados.get('volume_negocios', 0)
        ebitda = dados.get('ebitda', 0)

        autonomia = (capital / ativo * 100) if ativo > 0 else 0
        margem_ebitda = (ebitda / volume * 100) if volume > 0 else 0

        # Determinar rating
        if autonomia > 50 and margem_ebitda > 10:
            rating = "BAIXO"
        elif autonomia > 30 and margem_ebitda > 5:
            rating = "MÉDIO"
        elif autonomia > 20 and margem_ebitda > 0:
            rating = "ALTO"
        else:
            rating = "CRÍTICO"

        print(f"📊 Métricas calculadas:")
        print(f"   Autonomia Financeira: {autonomia:.1f}%")
        print(f"   Margem EBITDA: {margem_ebitda:.1f}%")
        print(f"   Rating: {rating}")

        # Gerar análise
        prompt = f"""Analise a saúde financeira desta empresa com base nos dados:

Dados Financeiros:
- Empresa: {dados.get('designacao_social')}
- Ativo Total: €{ativo:,.2f}
- Capital Próprio: €{capital:,.2f}
- Volume Negócios: €{volume:,.2f}
- EBITDA: €{ebitda:,.2f}

Métricas:
- Autonomia Financeira: {autonomia:.1f}%
- Margem EBITDA: {margem_ebitda:.1f}%

Forneça:
1. Nível de risco (BAIXO/MÉDIO/ALTO/CRÍTICO)
2. 3 pontos fortes
3. 3 pontos fracos/riscos
4. 3 recomendações para melhoria

Seja conciso e objetivo. Formato JSON."""

        message = client.messages.create(
            model="claude-3-5-sonnet-20241022",  # Usando Sonnet para teste
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}]
        )

        analise = message.content[0].text
        print(f"\n📝 Análise gerada:")
        print(analise)

        # Salvar análise
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        analise_file = OUTPUT_DIR / f"analise_{timestamp}.json"

        analise_data = {
            "empresa": dados.get('designacao_social'),
            "nif": dados.get('nif'),
            "ano": dados.get('ano_exercicio'),
            "metricas": {
                "autonomia_financeira": autonomia,
                "margem_ebitda": margem_ebitda,
                "rating": rating
            },
            "analise": analise,
            "dados_financeiros": dados
        }

        with open(analise_file, 'w', encoding='utf-8') as f:
            json.dump(analise_data, f, ensure_ascii=False, indent=2)

        print(f"\n✅ Análise salva: {analise_file}")

    except Exception as e:
        print(f"❌ Erro na análise: {e}")
        import traceback
        traceback.print_exc()

def main():
    """Função principal"""
    print("=" * 60)
    print("🧪 TESTE REAL - AUTOFUND AI")
    print("=" * 60)

    # 1. Extração do PDF
    dados = test_pdf_extraction()

    if dados:
        # 2. Geração Excel
        excel_file = test_excel_generation(dados)

        # 3. Análise Financeira
        test_financial_analysis(dados)

        print("\n" + "=" * 60)
        print("✅ TESTE CONCLUÍDO COM SUCESSO!")
        print(f"📊 Resultados salvos em: {OUTPUT_DIR}")
        print("=" * 60)
    else:
        print("\n❌ Falha no teste. Verifique os erros acima.")

if __name__ == "__main__":
    main()