#!/usr/bin/env python3
"""
AutoFund AI - POC Versão 3 Corrigida
Processa IES PDF → Extração Claude 3.5 → Validação → Análise Opus → Excel IAPMEI

Correções implementadas:
- Tratamento correto de células merged no Excel
- Melhoria na busca de células adjacente
- Verificação de tipo de célula antes de escrita
"""

import os
import json
import logging
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple
from dataclasses import dataclass
from datetime import datetime
import re

# Bibliotecas principais
import anthropic
import pandas as pd
import numpy as np
from pydantic import BaseModel, Field, validator, model_validator
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from dotenv import load_dotenv

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('autofund_ai.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Carregar variáveis de ambiente
load_dotenv()

# Configuração
ANTHROPIC_API_KEY = os.getenv('ANTHROPIC_API_KEY', 'sua-api-key-aqui')
IES_PDF_PATH = "IES - 2023.pdf"
TEMPLATE_PATH = "template_iapmei.xlsx"
OUTPUT_DIR = Path("outputs")
OUTPUT_DIR.mkdir(exist_ok=True)

# Cores para formatação Excel
COLOR_RED = "FFFF0000"
COLOR_YELLOW = "FFFFFF00"
COLOR_GREEN = "FF00FF00"


class ExtracoesFinanceiras(BaseModel):
    """Modelo Pydantic robusto para dados financeiros extraídos do IES"""

    # Identificação
    nome_empresa: str = Field(..., description="Nome da empresa")
    nif: str = Field(..., description="Número de Identificação Fiscal")
    periodo: str = Field(..., description="Ano fiscal")
    cae: Optional[str] = Field(None, description="Código de Atividade Económica")

    # Demonstração de Resultados (Valores em €)
    volume_negocios: float = Field(..., ge=0, description="Volume de Negócios (Vendas + Prestações)")
    custo_mercadorias: float = Field(0, ge=0, description="Custo das Mercadorias Vendidas")
    custo_materias: float = Field(0, ge=0, description="Custo das Matérias-primas")
    fornecimento_servicos: float = Field(0, ge=0, description="Fornecimento e Serviços Externos")
    custos_pessoal: float = Field(0, ge=0, description="Gastos com o Pessoal")
    depreciacoes: float = Field(0, ge=0, description="Gastos de Depreciação e de Amortização")
    ebitda: Optional[float] = Field(None, description="EBITDA calculado")
    resultados_operacionais: float = Field(..., description="Resultados Operacionais")
    resultados_financeiros: float = Field(0, description="Resultados Financeiros")
    resultados_antes_imposto: float = Field(..., description="Resultados Antes de Imposto")
    imposto_periodo: float = Field(0, ge=0, description="Imposto sobre o Rendimento do Período")
    resultado_liquido: float = Field(..., description="Resultado Líquido do Período")

    # Balanço (Valores em €)
    ativo_corrente: float = Field(..., ge=0, description="Ativo Corrente")
    ativo_nao_corrente: float = Field(..., ge=0, description="Ativo Não Corrente")
    total_ativo: float = Field(..., ge=0, description="Total do Ativo")

    passivo_corrente: float = Field(..., ge=0, description="Passivo Corrente")
    passivo_nao_corrente: float = Field(..., ge=0, description="Passivo Não Corrente")
    total_passivo: float = Field(..., ge=0, description="Total do Passivo")

    capital_proprio: float = Field(..., description="Capital Próprio")

    # Indicadores adicionais
    total_ativo_bruto: Optional[float] = Field(None, ge=0)
    total_amortizacoes_acumuladas: Optional[float] = Field(None)
    investimentos_em_imobilizado: Optional[float] = Field(None)

    @validator('nif')
    def validate_nif(cls, v):
        """Valida formato do NIF português"""
        if not re.match(r'^\d{9}$', v.replace(' ', '')):
            raise ValueError('NIF deve ter 9 dígitos')
        return v.replace(' ', '')

    @validator('ebitda', always=True)
    def calculate_ebitda(cls, v, values):
        """Calcula EBITDA se não fornecido"""
        if v is not None:
            return v
        # EBITDA = Resultado Operacional + Depreciações
        return values.get('resultados_operacionais', 0) + values.get('depreciacoes', 0)

    @model_validator(mode='after')
    def validate_accounting_equation(self):
        """Valida equação contabilística fundamental: Ativo = Passivo + Capital Próprio"""
        total_ativo = self.total_ativo
        total_passivo = self.total_passivo
        capital_proprio = self.capital_proprio

        # Permitir pequena diferença devido a arredondamentos
        diferenca = abs(total_ativo - (total_passivo + capital_proprio))
        if diferenca > 100:  # Tolerância de 100€
            logger.warning(f"Equação contabilística não bate: Ativo({total_ativo}) ≠ Passivo({total_passivo}) + CP({capital_proprio})")
            object.__setattr__(self, '_contabilidade_bate', False)
        else:
            object.__setattr__(self, '_contabilidade_bate', True)

        return self

    class Config:
        extra = "forbid"  # Não permite campos adicionais não especificados


class AnaliseFinanceira(BaseModel):
    """Modelo para análise de risco e recomendações"""

    # Indicadores calculados
    autonomia_financeira: float = Field(..., description="Capital Próprio / Total Ativo")
    liquidez_geral: float = Field(..., description="Ativo Corrente / Passivo Corrente")
    margem_ebitda: float = Field(..., description="EBITDA / Volume Negócios")
    rentabilidade_ativos: float = Field(..., description="Resultado Líquido / Total Ativo")
    endividamento: float = Field(..., description="Total Passivo / Total Ativo")

    # Classificação de risco
    nivel_risco: str = Field(..., description="BAIXO | MÉDIO | ALTO | CRÍTICO")

    # Análise qualitativa
    pontos_fortes: List[str] = Field(default_factory=list)
    pontos_fracos: List[str] = Field(default_factory=list)
    recomendacoes: List[str] = Field(default_factory=list)

    # Texto para candidatura
    memoria_descritiva: str = Field(..., description="Texto justificativo para IAPMEI")

    class Config:
        extra = "forbid"


class DataExtractor:
    """Classe responsável pela extração de dados do PDF IES usando Claude 3.5 Sonnet"""

    def __init__(self, api_key: str):
        # Handle custom base URL if configured
        base_url = os.getenv('ANTHROPIC_BASE_URL')
        auth_token = os.getenv('ANTHROPIC_AUTH_TOKEN')

        # Check if we're using a custom proxy/API gateway
        if base_url and auth_token:
            self.client = anthropic.Anthropic(
                api_key=auth_token,
                base_url=base_url
            )
        else:
            self.client = anthropic.Anthropic(api_key=api_key)
        self.file_id = None  # Initialize file ID

    def upload_pdf(self, pdf_path: str) -> str:
        """Faz upload do PDF IES para a Anthropic Files API"""
        try:
            with open(pdf_path, "rb") as f:
                response = self.client.beta.files.upload(
                    file=(os.path.basename(pdf_path), f, "application/pdf"),
                    purpose="assistants"
                )
            self.file_id = response.id
            logger.info(f"PDF uploaded com file_id: {self.file_id}")
            return self.file_id
        except Exception as e:
            logger.error(f"Erro ao fazer upload do PDF: {str(e)}")
            raise

    def extract_financial_data(self) -> Dict[str, Any]:
        """Extrai dados financeiros estruturados do IES"""
        if not self.file_id:
            raise ValueError("PDF não foi uploaded")

        prompt = """
        EXTRAI dados financeiros do IES (Informação Empresarial Simplificada) de Portugal.

        INSTRUÇÕES CRÍTICAS:
        1. Se uma tabela continuar na página seguinte (ex: Quadro 03-A), mantém o contexto
        2. Valores em EUR (€) - ignora "Milhares de Euros"
        3. Se não encontrares um valor específico, usa 0
        4. NÃO faças cálculos, apenas extrai valores diretos
        5. Verifica subtotais quando disponíveis

        CAMPOS OBRIGATÓRIOS:
        - Nome da empresa e NIF (geralmente página 1)
        - Ano fiscal
        - CAE (se disponível)

        QUADRO 03-A (Demonstração de Resultados):
        - Volume de Negócios (linha 71/72)
        - Custo das Mercadorias (linha 61)
        - Fornecimento e Serviços Externos (linha 62)
        - Gastos com o Pessoal (linha 64)
        - Gastos de Depreciação e Amortização (linha 64A)
        - Resultados Operacionais (antes de juros e impostos)
        - Resultados Financeiros
        - Resultado Antes de Imposto
        - Imposto sobre o Rendimento
        - Resultado Líquido

        QUADRO 04-A (Balanço):
        ATIVO:
        - Ativo Corrente (total)
        - Ativo Não Corrente (total)
        - Total do Ativo

        PASSIVO E CAPITAL PRÓPRIO:
        - Passivo Corrente (total)
        - Passivo Não Corrente (total)
        - Total do Passivo
        - Capital Próprio (total)

        RETORNA JSON válido:
        {
          "nome_empresa": "...",
          "nif": "...",
          "periodo": "2023",
          "cae": "...",
          "volume_negocios": 0.0,
          "custo_mercadorias": 0.0,
          "custo_materias": 0.0,
          "fornecimento_servicos": 0.0,
          "custos_pessoal": 0.0,
          "depreciacoes": 0.0,
          "resultados_operacionais": 0.0,
          "resultados_financeiros": 0.0,
          "resultados_antes_imposto": 0.0,
          "imposto_periodo": 0.0,
          "resultado_liquido": 0.0,
          "ativo_corrente": 0.0,
          "ativo_nao_corrente": 0.0,
          "total_ativo": 0.0,
          "passivo_corrente": 0.0,
          "passivo_nao_corrente": 0.0,
          "total_passivo": 0.0,
          "capital_proprio": 0.0
        }
        """

        try:
            message = self.client.beta.messages.create(
                model="claude-3-5-sonnet-20241022",
                max_tokens=4000,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "document",
                                "source": {
                                    "type": "file",
                                    "file_id": self.file_id,
                                },
                            },
                            {
                                "type": "text",
                                "text": prompt
                            }
                        ],
                    }
                ],
                betas=["pdf-to-structured-json-2024-04-01"],
            )

            # Parse do JSON
            json_str = message.content[0].text
            # Limpar possíveis marcas de código
            json_str = re.sub(r'```json\n?|\n?```', '', json_str).strip()

            data = json.loads(json_str)
            logger.info("Dados extraídos com sucesso")
            return data

        except json.JSONDecodeError as e:
            logger.error(f"Erro ao parsear JSON: {e}")
            logger.error(f"Resposta bruta: {message.content[0].text}")
            raise
        except Exception as e:
            logger.error(f"Erro na extração: {str(e)}")
            raise


class FinancialAnalyzer:
    """Classe para análise financeira e geração de insights usando Claude Opus 4.5"""

    def __init__(self, api_key: str):
        # Handle custom base URL if configured
        base_url = os.getenv('ANTHROPIC_BASE_URL')
        auth_token = os.getenv('ANTHROPIC_AUTH_TOKEN')

        # Check if we're using a custom proxy/API gateway
        if base_url and auth_token:
            self.client = anthropic.Anthropic(
                api_key=auth_token,
                base_url=base_url
            )
        else:
            self.client = anthropic.Anthropic(api_key=api_key)

    def calculate_ratios(self, data: ExtracoesFinanceiras) -> Dict[str, float]:
        """Calcula rácios financeiros"""
        ratios = {}

        # Autonomia Financeira
        ratios['autonomia_financeira'] = data.capital_proprio / data.total_ativo if data.total_ativo > 0 else 0

        # Liquidez Geral
        ratios['liquidez_geral'] = data.ativo_corrente / data.passivo_corrente if data.passivo_corrente > 0 else 0

        # Margem EBITDA
        ratios['margem_ebitda'] = data.ebitda / data.volume_negocios if data.volume_negocios > 0 else 0

        # Rentabilidade dos Ativos (ROA)
        ratios['rentabilidade_ativos'] = data.resultado_liquido / data.total_ativo if data.total_ativo > 0 else 0

        # Endividamento
        ratios['endividamento'] = data.total_passivo / data.total_ativo if data.total_ativo > 0 else 0

        return ratios

    def assess_risk_level(self, ratios: Dict[str, float]) -> str:
        """Classifica o nível de risco da empresa"""
        risk_score = 0

        # Autonomia Financeira (< 20% = risco)
        if ratios['autonomia_financeira'] < 0.20:
            risk_score += 3
        elif ratios['autonomia_financeira'] < 0.30:
            risk_score += 2
        elif ratios['autonomia_financeira'] < 0.40:
            risk_score += 1

        # Liquidez (< 1 = risco)
        if ratios['liquidez_geral'] < 1:
            risk_score += 3
        elif ratios['liquidez_geral'] < 1.5:
            risk_score += 1

        # Margem EBITDA (< 5% = risco)
        if ratios['margem_ebitda'] < 0.05:
            risk_score += 2
        elif ratios['margem_ebitda'] < 0.10:
            risk_score += 1

        # Rentabilidade negativa
        if ratios['rentabilidade_ativos'] < 0:
            risk_score += 3

        # Classificação final
        if risk_score >= 7:
            return "CRÍTICO"
        elif risk_score >= 5:
            return "ALTO"
        elif risk_score >= 3:
            return "MÉDIO"
        else:
            return "BAIXO"

    def generate_analysis(self, data: ExtracoesFinanceiras, context: str = "") -> AnaliseFinanceira:
        """Gera análise completa usando Claude Opus 4.5"""

        # Calcular rácios primeiro
        ratios = self.calculate_ratios(data)
        risk_level = self.assess_risk_level(ratios)

        # Preparar dados para o Opus
        financial_summary = {
            "empresa": data.nome_empresa,
            "nif": data.nif,
            "periodo": data.periodo,
            "volume_negocios": data.volume_negocios,
            "ebitda": data.ebitda,
            "resultado_liquido": data.resultado_liquido,
            "total_ativo": data.total_ativo,
            "capital_proprio": data.capital_proprio,
            "contabilidade_valida": data._contabilidade_bate,
            "ratios": ratios,
            "risco": risk_level
        }

        system_prompt = """
        És um consultor financeiro sénior especializado em candidaturas ao Portugal 2030/IAPMEI.

        A tua análise deve ser:
        1. OBJETIVA: Baseada apenas nos números fornecidos
        2. ESTRATÉGICA: Focada em what matters para aprovação
        3. CONSTRUTIVA: Cada problema deve ter uma solução
        4. FORMAL: Linguagem adequada para relatórios de candidatura

        Estrutura da resposta:
        - 3 pontos fortes (se aplicável)
        - 3 pontos fracos (sempre)
        - 3 recomendações acionáveis
        - Memória Descritiva (máx 400 palavras)
        """

        user_prompt = f"""
        Analisa esta empresa para candidatura Portugal 2030:

        DADOS FINANCEIROS:
        {json.dumps(financial_summary, indent=2)}

        CONTEXTO ADICIONAL (se fornecido):
        {context if context else "[Sem contexto adicional]"}

        INSTRUÇÕES ESPECÍFICAS:
        1. Se o Resultado Líquido for negativo, justifica como situação "conjuntural e não estrutural"
        2. Se a Autonomia Financeira for baixa (<30%), sugere capitalização ou conversão de suprimentos
        3. Se a Liquidez for <1.5, alerta para risco de solvabilidade a curto prazo
        4. Se o EBITDA for positivo mas RL negativo, explica impacto de custos não recorrentes

        A Memória Descritiva deve:
        - Começar com enquadramento positivo
        - Abordar honestamente os desafios
        - Apresentar plano de melhorias
        - Terminar com visão otimista mas realista

        Retorna JSON válido:
        {{
          "pontos_fortes": ["ponto1", "ponto2", "ponto3"],
          "pontos_fracos": ["ponto1", "ponto2", "ponto3"],
          "recomendacoes": ["rec1", "rec2", "rec3"],
          "memoria_descritiva": "Texto completo..."
        }}
        """

        try:
            response = self.client.messages.create(
                model="claude-opus-4-20250514",
                max_tokens=4000,
                system=system_prompt,
                messages=[
                    {
                        "role": "user",
                        "content": user_prompt
                    }
                ],
                temperature=0.3  # Mais consistente para análises
            )

            # Parse da resposta
            analysis_text = response.content[0].text
            analysis_text = re.sub(r'```json\n?|\n?```', '', analysis_text).strip()
            analysis_data = json.loads(analysis_text)

            # Criar objeto AnaliseFinanceira
            return AnaliseFinanceira(
                autonomia_financeira=ratios['autonomia_financeira'],
                liquidez_geral=ratios['liquidez_geral'],
                margem_ebitda=ratios['margem_ebitda'],
                rentabilidade_ativos=ratios['rentabilidade_ativos'],
                endividamento=ratios['endividamento'],
                nivel_risco=risk_level,
                pontos_fortes=analysis_data.get('pontos_fortes', []),
                pontos_fracos=analysis_data.get('pontos_fracos', []),
                recomendacoes=analysis_data.get('recomendacoes', []),
                memoria_descritiva=analysis_data.get('memoria_descritiva', '')
            )

        except Exception as e:
            logger.error(f"Erro na análise Opus: {str(e)}")
            # Fallback para análise básica
            return self._generate_fallback_analysis(data, ratios, risk_level)

    def _generate_fallback_analysis(self, data: ExtracoesFinanceiras, ratios: Dict[str, float], risk_level: str) -> AnaliseFinanceira:
        """Gera análise básica sem depender do Opus"""

        pontos_fracos = []
        recomendacoes = []

        if ratios['autonomia_financeira'] < 0.30:
            pontos_fracos.append("Autonomia financeira reduzida")
            recomendacoes.append("Considerar aumento de capital social")

        if ratios['liquidez_geral'] < 1.5:
            pontos_fracos.append("Liquidez restrita")
            recomendacoes.append("Renegociar prazos com fornecedores")

        if data.resultado_liquido < 0:
            pontos_fracos.append("Resultado negativo no período")
            recomendacoes.append("Implementar plano de recuperação")

        memoria = f"""
        A empresa {data.nome_empresa} apresenta um volume de negócios de €{data.volume_negocios:,.2f}
        para o período de {data.periodo}.

        Os indicadores financeiros demonstram {risk_level.lower()} nível de risco,
        com uma autonomia financeira de {ratios['autonomia_financeira']:.1%}.

        {"Apesar dos desafios enfrentados, a empresa mantém capacidade de recuperação." if risk_level != "BAIXO" else "A empresa demonstra sólida posição financeira."}

        Projetamos para os próximos períodos uma melhoria gradual dos indicadores,
        suportada pelas medidas de reestruturação implementadas.
        """

        return AnaliseFinanceira(
            autonomia_financeira=ratios['autonomia_financeira'],
            liquidez_geral=ratios['liquidez_geral'],
            margem_ebitda=ratios['margem_ebitda'],
            rentabilidade_ativos=ratios['rentabilidade_ativos'],
            endividamento=ratios['endividamento'],
            nivel_risco=risk_level,
            pontos_fortes=[f"Volume de negócios: €{data.volume_negocios:,.2f}"],
            pontos_fracos=pontos_fracos,
            recomendacoes=recomendacoes,
            memoria_descritiva=memoria.strip()
        )


class ExcelGenerator:
    """Classe para preenchimento do template Excel IAPMEI"""

    def __init__(self, template_path: str):
        self.template_path = template_path

        # Mapeamento de labels para procurar no Excel
        self.field_mappings = {
            "nome_empresa": ["Nome da empresa", "Empresa", "Designação"],
            "nif": ["NIF", "Nº de Identificação Fiscal", "Contribuinte"],
            "periodo": ["Ano", "Período", "Exercício"],
            "volume_negocios": ["Volume de Negócios", "Vendas", "Total Vendas"],
            "ebitda": ["EBITDA", "Resultado antes de depreciações", "RAAJI"],
            "resultado_liquido": ["Resultado Líquido", "Resultado do Período"],
            "total_ativo": ["Total do Ativo", "Ativo Total"],
            "capital_proprio": ["Capital Próprio", "Situação Líquida"],
            "autonomia_financeira": ["Autonomia Financeira", "Autonomia (%)", "AF (%)"],
            "liquidez_geral": ["Liquidez Geral", "Liquidez"],
            "margem_ebitda": ["Margem EBITDA", "EBITDA %"],
            "nivel_risco": ["Nível de Risco", "Risco"],
            "memoria_descritiva": ["Memória Descritiva", "Descrição", "Justificação"]
        }

    def find_cell_by_label(self, ws, label: str, search_direction: str = "right") -> Optional[Tuple[int, int]]:
        """Procura por um label na worksheet e retorna célula adjacente livre"""
        label_lower = label.lower()

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_value = cell.value.lower().strip()
                    # Match exato ou parcial
                    if label_lower in cell_value:
                        # Retorna célula adjacente
                        row_idx, col_idx = cell.row, cell.column

                        if search_direction == "right":
                            # Procura próxima célula vazia à direita
                            for c in range(col_idx + 1, min(col_idx + 5, ws.max_column + 1)):
                                target_cell = ws.cell(row=row_idx, column=c)
                                # Verifica se célula não está em merge range
                                if not self._is_cell_merged(ws, target_cell):
                                    return (row_idx, c)
                        elif search_direction == "down":
                            # Procura próxima célula vazia abaixo
                            for r in range(row_idx + 1, min(row_idx + 5, ws.max_row + 1)):
                                target_cell = ws.cell(row=r, column=col_idx)
                                if not self._is_cell_merged(ws, target_cell):
                                    return (r, col_idx)

        return None

    def _is_cell_merged(self, ws, cell) -> bool:
        """Verifica se célula faz parte de um merged range"""
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return True
        return False

    def fill_template(self, data: ExtracoesFinanceiras, analysis: AnaliseFinanceira, output_path: str):
        """Preenche o template Excel com dados extraídos e análise"""

        try:
            # Carregar ou criar template
            if os.path.exists(self.template_path):
                wb = load_workbook(self.template_path)
            else:
                logger.warning(f"Template não encontrado em {self.template_path}. Criando novo.")
                wb = Workbook()
                ws = wb.active
                ws.title = "Análise Financeira"

                # Criar headers básicos
                headers = ["Campo", "Valor", "Análise"]
                ws.append(headers)

            # Usar primeira worksheet ou criar
            if not wb.worksheets:
                ws = wb.active
            else:
                ws = wb.worksheets[0]

            # Preparar dados para preencher
            all_data = {
                **data.model_dump(),
                **{
                    'autonomia_financeira': analysis.autonomia_financeira,
                    'liquidez_geral': analysis.liquidez_geral,
                    'margem_ebitda': analysis.margem_ebitda,
                    'nivel_risco': analysis.nivel_risco,
                    'memoria_descritiva': analysis.memoria_descritiva
                }
            }

            # Preencher dados usando mapeamento por labels
            for field, labels in self.field_mappings.items():
                if field in all_data:
                    value = all_data[field]

                    # Tentar encontrar célula para cada label possível
                    for label in labels:
                        cell_pos = self.find_cell_by_label(ws, label)
                        if cell_pos:
                            row, col = cell_pos
                            target_cell = ws.cell(row=row, column=col)

                            # Verifica se célula não está merged
                            if not self._is_cell_merged(ws, target_cell):
                                target_cell.value = value

                            # Formatação condicional para indicadores de risco
                            if field in ['autonomia_financeira', 'liquidez_geral', 'margem_ebitda']:
                                if isinstance(value, (int, float)):
                                    if value < 0.2:
                                        fill = PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid")
                                    elif value < 0.3:
                                        fill = PatternFill(start_color=COLOR_YELLOW, end_color=COLOR_YELLOW, fill_type="solid")
                                    else:
                                        fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
                                    target_cell.fill = fill

                            # Formatar nível de risco
                            elif field == 'nivel_risco':
                                if value == "CRÍTICO":
                                    fill = PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type="solid")
                                elif value == "ALTO":
                                    fill = PatternFill(start_color=COLOR_YELLOW, end_color=COLOR_YELLOW, fill_type="solid")
                                else:
                                    fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
                                target_cell.fill = fill

                            break  # Usa primeira label encontrada

            # Adicionar resumo da análise em nova sheet se não existir
            if "Resumo" not in wb.sheetnames:
                resumo_ws = wb.create_sheet("Resumo")

                # Título
                resumo_ws['A1'] = "Resumo da Análise - AutoFund AI"
                resumo_ws['A1'].font = Font(bold=True, size=14)

                # Informações básicas
                resumo_ws['A3'] = f"Empresa: {data.nome_empresa}"
                resumo_ws['A4'] = f"NIF: {data.nif}"
                resumo_ws['A5'] = f"Período: {data.periodo}"
                resumo_ws['A6'] = f"Nível de Risco: {analysis.nivel_risco}"

                # Pontos fortes e fracos
                row = 8
                resumo_ws[f'A{row}'] = "Pontos Fortes:"
                resumo_ws[f'A{row}'].font = Font(bold=True)
                for i, ponto in enumerate(analysis.pontos_fortes, 1):
                    resumo_ws[f'A{row+i}'] = f"• {ponto}"

                row += len(analysis.pontos_fortes) + 2
                resumo_ws[f'A{row}'] = "Pontos Fracos:"
                resumo_ws[f'A{row}'].font = Font(bold=True)
                for i, ponto in enumerate(analysis.pontos_fracos, 1):
                    resumo_ws[f'A{row+i}'] = f"• {ponto}"

                row += len(analysis.pontos_fracos) + 2
                resumo_ws[f'A{row}'] = "Recomendações:"
                resumo_ws[f'A{row}'].font = Font(bold=True)
                for i, rec in enumerate(analysis.recomendacoes, 1):
                    resumo_ws[f'A{row+i}'] = f"{i}. {rec}"

                # Ajustar largura das colunas
                for column in resumo_ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    resumo_ws.column_dimensions[column_letter].width = adjusted_width

            # Salvar ficheiro
            wb.save(output_path)
            logger.info(f"Excel salvo em: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Erro ao preencher Excel: {str(e)}")
            raise


class AutoFundAI:
    """Classe principal orquestradora do pipeline"""

    def __init__(self, api_key: str):
        self.api_key = api_key
        self.extractor = DataExtractor(api_key)
        self.analyzer = FinancialAnalyzer(api_key)
        self.excel_generator = ExcelGenerator(TEMPLATE_PATH)

    def process_ies(self, pdf_path: str, context: str = "") -> Dict[str, Any]:
        """Pipeline completo de processamento do IES"""

        results = {}

        try:
            # 1. Upload e extração
            logger.info("Iniciando upload e extração do IES...")
            file_id = self.extractor.upload_pdf(pdf_path)

            # 2. Extrair dados financeiros
            logger.info("Extraindo dados financeiros...")
            raw_data = self.extractor.extract_financial_data()

            # 3. Validar com Pydantic
            logger.info("Validando dados extraídos...")
            financial_data = ExtracoesFinanceiras(**raw_data)
            logger.info(f"Validação: Contabilidade bate? {financial_data._contabilidade_bate}")

            # 4. Análise com Opus
            logger.info("Gerando análise financeira...")
            analysis = self.analyzer.generate_analysis(financial_data, context)

            # 5. Gerar Excel
            logger.info("Preenchendo template Excel...")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_filename = f"autofund_analysis_{financial_data.nif}_{timestamp}.xlsx"
            excel_path = OUTPUT_DIR / excel_filename

            self.excel_generator.fill_template(financial_data, analysis, str(excel_path))

            # 6. Gerar relatório JSON
            report = {
                "metadata": {
                    "empresa": financial_data.nome_empresa,
                    "nif": financial_data.nif,
                    "periodo": financial_data.periodo,
                    "data_processamento": datetime.now().isoformat(),
                    "versao": "1.0.0"
                },
                "dados_financeiros": financial_data.model_dump(),
                "analise": analysis.model_dump(),
                "ficheiros_gerados": {
                    "excel": str(excel_path),
                    "json": str(OUTPUT_DIR / f"analysis_{financial_data.nif}_{timestamp}.json")
                }
            }

            # 7. Salvar relatório JSON
            json_path = OUTPUT_DIR / f"analysis_{financial_data.nif}_{timestamp}.json"
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(report, f, indent=2, ensure_ascii=False)

            logger.info(f"Processo concluído com sucesso!")
            logger.info(f"Excel: {excel_path}")
            logger.info(f"JSON: {json_path}")

            return report

        except Exception as e:
            logger.error(f"Erro no processamento: {str(e)}")
            raise


def main():
    """Função principal para execução local"""

    # Verificar API key
    if ANTHROPIC_API_KEY == 'sua-api-key-aqui':
        print("❌ ERRO: Configure sua API Key da Anthropic no arquivo .env")
        print("Exemplo: ANTHROPIC_API_KEY=sk-ant-xxx")
        return

    # Verificar PDF
    if not os.path.exists(IES_PDF_PATH):
        print(f"❌ ERRO: PDF IES não encontrado em {IES_PDF_PATH}")
        return

    # Contexto adicional (opcional)
    context_exemplo = """
    A empresa enfrentou desafios pós-COVID com queda no setor de construção civil em 2023.
    No entanto, conseguiu contratos importantes para 2024 e está a investir em digitalização.
    Tem planos de expandir para mercados internacionais.
    """

    print("🚀 Iniciando AutoFund AI...")
    print(f"📄 Processando: {IES_PDF_PATH}")

    # Inicializar e processar
    autofund = AutoFundAI(ANTHROPIC_API_KEY)

    try:
        resultado = autofund.process_ies(IES_PDF_PATH, context_exemplo)

        print("\n✅ RESULTADO:")
        print(json.dumps(resultado["metadata"], indent=2))
        print(f"\n📊 Nível de Risco: {resultado['analise']['nivel_risco']}")
        print(f"💪 Pontos Fortes: {len(resultado['analise']['pontos_fortes'])}")
        print(f"⚠️ Pontos Fracos: {len(resultado['analise']['pontos_fracos'])}")
        print(f"💡 Recomendações: {len(resultado['analise']['recomendacoes'])}")

    except Exception as e:
        print(f"\n❌ Erro: {str(e)}")
        logger.exception("Detalhes do erro:")


if __name__ == "__main__":
    main()