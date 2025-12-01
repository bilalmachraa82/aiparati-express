#!/usr/bin/env python3
"""
Test Real IES PDF Processing with AutoFund AI
Tests the complete workflow with actual IES file
"""

import requests
import json
import time
import os
from pathlib import Path
from datetime import datetime

class TestRealIESProcessing:
    """Test IES processing with real files"""

    def __init__(self):
        self.api_base = "http://localhost:8000"
        self.auth_token = "test-token-12345"
        self.headers = {
            "Authorization": f"Bearer {self.auth_token}",
            "Accept": "application/json"
        }

    def test_01_health_check(self):
        """Test API is running"""
        print("🔍 Testing API Health...")

        response = requests.get(f"{self.api_base}/health", timeout=5)

        assert response.status_code == 200, f"Health check failed: {response.status_code}"

        data = response.json()
        assert data["status"] == "healthy"
        assert "version" in data

        print(f"✅ API is healthy - Version: {data['version']}")
        return True

    def test_02_upload_real_ies(self):
        """Upload real IES PDF for processing"""
        print("📄 Testing Real IES Upload...")

        ies_path = Path("IES - 2023.pdf")
        if not ies_path.exists():
            print("❌ IES - 2023.pdf not found")
            return False

        form_data = {
            "nif": "516807706",
            "ano_exercicio": "2023",
            "designacao_social": "PLF - PROJETOS, LDA.",
            "email": "test@plf.pt"
        }

        with open(ies_path, "rb") as f:
            files = {"file": ("IES - 2023.pdf", f, "application/pdf")}
            response = requests.post(
                f"{self.api_base}/api/upload",
                files=files,
                data=form_data,
                headers=self.headers,
                timeout=30
            )

        print(f"📊 Upload response status: {response.status_code}")

        if response.status_code != 200:
            print(f"❌ Upload failed: {response.text}")
            return False

        data = response.json()
        assert "task_id" in data
        assert data["status"] == "processing"

        task_id = data["task_id"]
        print(f"✅ Upload successful - Task ID: {task_id}")
        return task_id

    def test_03_monitor_processing(self, task_id):
        """Monitor IES processing progress"""
        print("⏳ Monitoring Processing Progress...")

        max_wait_time = 300  # 5 minutes max
        start_time = time.time()

        while time.time() - start_time < max_wait_time:
            response = requests.get(
                f"{self.api_base}/api/status/{task_id}",
                headers=self.headers,
                timeout=10
            )

            assert response.status_code == 200, f"Status check failed: {response.status_code}"

            data = response.json()
            status = data["status"]
            print(f"🔄 Status: {status}")

            if status == "completed":
                print("✅ Processing completed successfully!")
                return data
            elif status == "error":
                print(f"❌ Processing failed: {data.get('error', 'Unknown error')}")
                return None

            time.sleep(5)  # Wait 5 seconds before checking again

        print("⏰ Processing timeout")
        return None

    def test_04_validate_results(self, task_data):
        """Validate processing results"""
        print("🔍 Validating Processing Results...")

        if not task_data or "result" not in task_data:
            print("❌ No results to validate")
            return False

        result = task_data["result"]

        # Check financial data
        assert "dados_financeiros" in result
        financial_data = result["dados_financeiros"]

        # Check analysis
        assert "analise" in result
        analysis = result["analise"]

        # Check metadata
        assert "metadata" in result
        metadata = result["metadata"]

        # Validate key fields
        assert "nif" in metadata
        assert "ano_exercicio" in metadata
        assert "designacao_social" in metadata
        assert metadata["nif"] == "516807706"
        assert metadata["ano_exercicio"] == "2023"

        print(f"✅ Company: {metadata['designacao_social']}")
        print(f"✅ NIF: {metadata['nif']}")
        print(f"✅ Year: {metadata['ano_exercicio']}")

        # Validate financial metrics
        if "volume_negocios" in financial_data:
            print(f"✅ Volume de Negócios: €{financial_data['volume_negocios']:,.2f}")

        if "ebitda" in financial_data:
            print(f"✅ EBITDA: €{financial_data['ebitda']:,.2f}")

        # Validate risk rating
        if "rating" in analysis:
            print(f"✅ Risk Rating: {analysis['rating']}")

        return True

    def test_05_download_generated_files(self, task_id):
        """Test download of generated files"""
        print("📥 Testing File Downloads...")

        # Test Excel download
        excel_response = requests.get(
            f"{self.api_base}/api/download/{task_id}/excel",
            headers=self.headers,
            timeout=30
        )

        if excel_response.status_code == 200:
            excel_path = Path(f"outputs/test_downloaded_{task_id}.xlsx")
            excel_path.parent.mkdir(exist_ok=True)

            with open(excel_path, "wb") as f:
                f.write(excel_response.content)

            file_size = excel_path.stat().st_size
            print(f"✅ Excel downloaded successfully - Size: {file_size} bytes")
        else:
            print(f"❌ Excel download failed: {excel_response.status_code}")
            return False

        # Test JSON download
        json_response = requests.get(
            f"{self.api_base}/api/download/{task_id}/json",
            headers=self.headers,
            timeout=30
        )

        if json_response.status_code == 200:
            json_path = Path(f"outputs/test_downloaded_{task_id}.json")

            with open(json_path, "wb") as f:
                f.write(json_response.content)

            # Validate JSON content
            with open(json_path, "r") as f:
                json_data = json.load(f)

            print(f"✅ JSON downloaded successfully - Size: {len(json.dumps(json_data))} chars")
        else:
            print(f"❌ JSON download failed: {json_response.status_code}")
            return False

        return True

    def test_06_validate_excel_format(self, task_id):
        """Validate generated Excel format"""
        print("📊 Validating Excel Format...")

        excel_path = Path(f"outputs/test_downloaded_{task_id}.xlsx")
        if not excel_path.exists():
            print("❌ Excel file not found")
            return False

        try:
            import openpyxl

            # Load the workbook
            wb = openpyxl.load_workbook(excel_path)
            sheet_names = wb.sheetnames

            print(f"✅ Excel loaded successfully - Sheets: {sheet_names}")

            # Check if it has expected structure
            if len(sheet_names) > 0:
                sheet = wb[sheet_names[0]]

                # Check for IAPMEI template markers
                found_markers = []
                for row in sheet.iter_rows(max_row=10, values_only=True):
                    for cell in row:
                        if cell and isinstance(cell, str):
                            if "IAPMEI" in cell.upper():
                                found_markers.append("IAPMEI")
                            elif "PORTUGAL 2030" in cell.upper():
                                found_markers.append("PORTUGAL 2030")
                            elif "CANDIDATURA" in cell.upper():
                                found_markers.append("CANDIDATURA")

                if found_markers:
                    print(f"✅ Found template markers: {found_markers}")
                else:
                    print("⚠️  No template markers found in first 10 rows")

                print(f"✅ Sheet dimensions: {sheet.max_row} rows x {sheet.max_column} columns")

            return True

        except Exception as e:
            print(f"❌ Excel validation failed: {e}")
            return False

    def test_07_error_handling(self):
        """Test API error handling"""
        print("🚨 Testing Error Handling...")

        # Test invalid task ID
        response = requests.get(
            f"{self.api_base}/api/status/invalid-task-id",
            headers=self.headers,
            timeout=10
        )

        assert response.status_code == 404, "Expected 404 for invalid task"
        print("✅ Invalid task ID properly returns 404")

        # Test invalid file type
        invalid_files = {
            "file": ("test.txt", "This is not a PDF", "text/plain"),
            "nif": "123456789",
            "ano_exercicio": "2023",
            "designacao_social": "Test",
            "email": "test@example.com"
        }

        response = requests.post(
            f"{self.api_base}/api/upload",
            files={"file": ("test.txt", "This is not a PDF", "text/plain")},
            data=invalid_files,
            headers=self.headers,
            timeout=10
        )

        # Should reject non-PDF files
        assert response.status_code in [400, 422], "Expected rejection of non-PDF files"
        print("✅ Non-PDF files properly rejected")

        return True

    def run_full_test_suite(self):
        """Run complete test suite"""
        print("🚀 Starting AutoFund AI Real IES Test Suite")
        print("=" * 60)

        test_results = []

        try:
            # Test 1: Health Check
            if self.test_01_health_check():
                test_results.append(("Health Check", "✅ PASS"))
            else:
                test_results.append(("Health Check", "❌ FAIL"))
                return test_results

            # Test 2: Upload Real IES
            task_id = self.test_02_upload_real_ies()
            if task_id:
                test_results.append(("IES Upload", "✅ PASS"))
            else:
                test_results.append(("IES Upload", "❌ FAIL"))
                return test_results

            # Test 3: Monitor Processing
            task_data = self.test_03_monitor_processing(task_id)
            if task_data:
                test_results.append(("Processing Monitor", "✅ PASS"))
            else:
                test_results.append(("Processing Monitor", "❌ FAIL"))
                return test_results

            # Test 4: Validate Results
            if self.test_04_validate_results(task_data):
                test_results.append(("Results Validation", "✅ PASS"))
            else:
                test_results.append(("Results Validation", "❌ FAIL"))

            # Test 5: Download Files
            if self.test_05_download_generated_files(task_id):
                test_results.append(("File Downloads", "✅ PASS"))
            else:
                test_results.append(("File Downloads", "❌ FAIL"))

            # Test 6: Validate Excel
            if self.test_06_validate_excel_format(task_id):
                test_results.append(("Excel Format", "✅ PASS"))
            else:
                test_results.append(("Excel Format", "❌ FAIL"))

            # Test 7: Error Handling
            if self.test_07_error_handling():
                test_results.append(("Error Handling", "✅ PASS"))
            else:
                test_results.append(("Error Handling", "❌ FAIL"))

        except Exception as e:
            print(f"❌ Test suite failed with error: {e}")
            test_results.append(("Test Suite", f"❌ ERROR: {e}"))

        return test_results


def main():
    """Main test execution"""
    # Check if API is running
    try:
        response = requests.get("http://localhost:8000/health", timeout=5)
        if response.status_code != 200:
            print("❌ API is not responding correctly")
            return
    except requests.exceptions.RequestException:
        print("❌ Cannot connect to API at http://localhost:8000")
        print("Please start the API server first: python api/main.py")
        return

    # Run tests
    tester = TestRealIESProcessing()
    results = tester.run_full_test_suite()

    # Print results
    print("\n" + "=" * 60)
    print("📋 TEST RESULTS SUMMARY")
    print("=" * 60)

    passed = 0
    for test_name, result in results:
        print(f"{test_name:25} {result}")
        if "✅ PASS" in result:
            passed += 1

    print(f"\nResults: {passed}/{len(results)} tests passed")

    if passed == len(results):
        print("\n🎉 All tests passed! AutoFund AI is working correctly.")
        print("\n📝 Test files generated in outputs/ directory")
        print("✅ Ready for production use")
    else:
        print(f"\n⚠️  {len(results) - passed} test(s) failed. Check logs above.")

    # Save test report
    report = {
        "test_date": datetime.now().isoformat(),
        "total_tests": len(results),
        "passed_tests": passed,
        "results": [{"test": name, "status": result} for name, result in results]
    }

    with open("outputs/test_real_ies_report.json", "w") as f:
        json.dump(report, f, indent=2)

    print(f"\n📊 Test report saved: outputs/test_real_ies_report.json")


if __name__ == "__main__":
    main()