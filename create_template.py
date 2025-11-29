#!/usr/bin/env python3
"""
Cria um template Excel básico do IAPMEI para testes
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_iapmei_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidatura Portugal 2030"

    # Estilos
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=16, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # Título
    ws['A1'] = "FORMULÁRIO DE CANDIDATURA - PORTUGAL 2030"
    ws['A1'].font = title_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:E1')

    # Dados da Empresa
    row = 3
    ws[f'A{row}'] = "DADOS DA EMPRESA"
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:C{row}')

    row += 1
    ws[f'A{row}'] = "Nome da empresa:"
    ws[f'B{row}'] = ""

    row += 1
    ws[f'A{row}'] = "NIF:"
    ws[f'B{row}'] = ""

    row += 1
    ws[f'A{row}'] = "Período:"
    ws[f'B{row}'] = ""

    row += 1
    ws[f'A{row}'] = "CAE:"
    ws[f'B{row}'] = ""

    # Indicadores Financeiros
    row += 2
    ws[f'A{row}'] = "INDICADORES FINANCEIROS"
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:C{row}')

    row += 1
    indicadores = [
        ("Volume de Negócios", ""),
        ("EBITDA", ""),
        ("Resultado Líquido", ""),
        ("Total do Ativo", ""),
        ("Capital Próprio", ""),
        ("Autonomia Financeira", ""),
        ("Liquidez Geral", ""),
        ("Margem EBITDA", ""),
        ("Nível de Risco", "")
    ]

    for label, value in indicadores:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1

    # Análise e Recomendações
    row += 1
    ws[f'A{row}'] = "ANÁLISE E RECOMENDAÇÕES"
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:C{row}')

    row += 1
    ws[f'A{row}'] = "Pontos Fortes:"
    ws[f'B{row}'] = ""

    row += 1
    ws[f'A{row}'] = "Pontos Fracos:"
    ws[f'B{row}'] = ""

    row += 1
    ws[f'A{row}'] = "Recomendações:"
    ws[f'B{row}'] = ""

    # Memória Descritiva
    row += 2
    ws[f'A{row}'] = "MEMÓRIA DESCRITIVA"
    ws[f'A{row}'].font = header_font
    ws.merge_cells(f'A{row}:C{row}')

    row += 1
    ws[f'A{row}'] = "Memória Descritiva:"
    ws.merge_cells(f'A{row}:E{row+10}')
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 50

    # Aplicar bordas a tudo
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = border

    # Salvar
    wb.save("template_iapmei.xlsx")
    print("✅ Template IAPMEI criado: template_iapmei.xlsx")

if __name__ == "__main__":
    create_iapmei_template()