#!/usr/bin/env python3
"""
Verifica células mescladas do template Excel
"""

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

TEMPLATE_PATH = "template_iapmei.xlsx"

def verificar_template():
    """Verifica estrutura do template"""
    print("🔍 Verificando estrutura do template Excel...")
    print(f"Arquivo: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    print(f"\n📊 Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")

    # Verificar células mescladas
    print("\n🔗 Células mescladas:")
    for merge in ws.merged_cells.ranges:
        print(f"   {merge}")

    # Mapeamento correto (evitando células mescladas)
    print("\n✅ Mapeamento correto (células não mescladas):")
    mapeamento = {
        "B4": "Nome da empresa",  # Verificar se não está mesclada
        "B6": "NIF",  # Verificar
        "F4": "Ano",  # Verificar
        "D15": "Ativo Total",
        "D18": "Passivo Total",
        "D21": "Capital Próprio",
        "D24": "Volume Negócios",
        "D27": "EBITDA"
    }

    for celula, descricao in mapeamento.items():
        coord = coordinate_from_string(celula)
        linha = coord[1]
        coluna = column_index_from_string(coord[0])
        celula_obj = ws.cell(row=linha, column=coluna)

        # Verificar se está mesclada
        is_merged = False
        for merge in ws.merged_cells.ranges:
            if celula_obj.coordinate in merge:
                is_merged = True
                break

        status = "✅" if not is_merged else "❌ Mesclada"
        print(f"   {celula} ({descricao}): {status}")

        # Mostrar conteúdo atual
        if celula_obj.value:
            print(f"      Conteúdo: {celula_obj.value}")

    # Sugerir células alternativas
    print("\n💡 Sugestão de células alternativas:")
    alternativas = {
        "Nome da empresa": "B4 (se não mesclada) ou B5",
        "NIF": "B6 ou B7",
        "Ano": "F4 ou F5",
        "Ativo Total": "D15 ou D16",
        "Passivo Total": "D18 ou D19",
        "Capital Próprio": "D21 ou D22",
        "Volume Negócios": "D24 ou D25",
        "EBITDA": "D27 ou D28"
    }

    for campo, celula in alternativas.items():
        print(f"   {campo}: {celula}")

if __name__ == "__main__":
    verificar_template()